from flask import Blueprint, jsonify, request, current_app, send_file
from app.models import Score, db, Tournament, Member
import pandas as pd
import os
from werkzeug.utils import secure_filename
import traceback
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import uuid
import csv
from sqlalchemy import func
from io import BytesIO
from datetime import datetime

bp = Blueprint('scores', __name__)

@bp.route('/', methods=['GET'])
def get_scores():
    try:
        tournament_id = request.args.get('tournament_id', type=int)
        if not tournament_id:
            return jsonify({'error': '必須提供賽事ID'}), 400
            
        scores = Score.query.filter_by(tournament_id=tournament_id).all()
        return jsonify([score.to_dict() for score in scores])
    except Exception as e:
        current_app.logger.error(f"獲取成績時發生錯誤: {str(e)}")
        return jsonify({'error': str(e)}), 500

@bp.route('/scores/<int:id>', methods=['GET'])
def get_score(id):
    score = Score.query.get_or_404(id)
    return jsonify(score.to_dict())

@bp.route('/import/<int:tournament_id>', methods=['POST'])
def import_scores(tournament_id):
    temp_path = None
    try:
        if 'file' not in request.files:
            return jsonify({'error': '未找到上傳的檔案'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '未選擇檔案'}), 400
        
        # 檢查檔案類型
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': '不支援的檔案格式，請使用 Excel 檔案 (.xlsx 或 .xls)'}), 400
        
        # 保存原始檔案名
        original_filename = file.filename
        
        # 創建臨時目錄（如果不存在）
        upload_dir = os.path.join(current_app.instance_path, 'uploads')
        os.makedirs(upload_dir, exist_ok=True)
        
        # 使用原始檔案名保存檔案
        temp_path = os.path.join(upload_dir, original_filename)
        
        # 保存檔案
        file.save(temp_path)
        
        try:
            # 使用 pandas 讀取 Excel 檔案
            df = pd.read_excel(temp_path)
            
            # 清理列名（移除空白和特殊字符）
            df.columns = df.columns.str.strip().str.replace('\ufeff', '').str.replace('\u3000', ' ')
            
            # 檢查必要的欄位是否存在
            required_columns = ['會員編號', 'HOLE', '姓名', '淨桿名次', '總桿數', 
                              '前次差點', '淨桿桿數', '差點增減', '新差點', '積分']
            
            # 檢查每個必要欄位
            missing_columns = []
            column_mapping = {}
            for col in required_columns:
                found = False
                for existing_col in df.columns:
                    if col in existing_col:  # 使用部分匹配
                        column_mapping[col] = existing_col
                        found = True
                        break
                if not found:
                    missing_columns.append(col)
            
            if missing_columns:
                raise Exception(f'Excel 檔案缺少以下欄位：{", ".join(missing_columns)}')
            
            # 重命名列以匹配所需的名稱
            df = df.rename(columns=column_mapping)
            
            # 清理數據
            for col in df.columns:
                if df[col].dtype == object:  # 如果是字符串類型
                    df[col] = df[col].astype(str).str.strip()
            
            # 刪除該賽事的所有現有成績
            Score.query.filter_by(tournament_id=tournament_id).delete()
            
            # 處理每一行數據
            for _, row in df.iterrows():
                try:
                    score = Score(tournament_id=tournament_id)
                    
                    # 設置各個欄位，處理可能的空值和特殊字符
                    score.member_number = str(row['會員編號']).strip() if pd.notna(row['會員編號']) else None
                    score.full_name = str(row['HOLE']).strip() if pd.notna(row['HOLE']) else None
                    score.chinese_name = str(row['姓名']).strip() if pd.notna(row['姓名']) else None
                    
                    # 處理數值欄位
                    try:
                        score.net_rank = int(row['淨桿名次']) if pd.notna(row['淨桿名次']) else None
                    except (ValueError, TypeError):
                        score.net_rank = None
                        
                    try:
                        score.gross_score = int(row['總桿數']) if pd.notna(row['總桿數']) else None
                    except (ValueError, TypeError):
                        score.gross_score = None
                        
                    try:
                        score.previous_handicap = float(row['前次差點']) if pd.notna(row['前次差點']) else None
                    except (ValueError, TypeError):
                        score.previous_handicap = None
                        
                    try:
                        score.net_score = float(row['淨桿桿數']) if pd.notna(row['淨桿桿數']) else None
                    except (ValueError, TypeError):
                        score.net_score = None
                        
                    try:
                        score.handicap_change = float(row['差點增減']) if pd.notna(row['差點增減']) else None
                    except (ValueError, TypeError):
                        score.handicap_change = None
                        
                    try:
                        score.new_handicap = float(row['新差點']) if pd.notna(row['新差點']) else None
                    except (ValueError, TypeError):
                        score.new_handicap = None
                        
                    try:
                        score.points = int(row['積分']) if pd.notna(row['積分']) else None
                    except (ValueError, TypeError):
                        score.points = None
                    
                    # 根據會員編號查找會員 ID
                    if score.member_number:
                        member = Member.query.filter_by(member_number=score.member_number).first()
                        if member:
                            score.member_id = member.id
                        else:
                            # 如果找不到會員，創建一個新的會員
                            member = Member(
                                member_number=score.member_number,
                                name=score.chinese_name,
                                handicap=score.new_handicap
                            )
                            db.session.add(member)
                            db.session.flush()  # 獲取新創建的會員 ID
                            score.member_id = member.id
                    
                    db.session.add(score)
                
                except Exception as row_error:
                    current_app.logger.error(f'處理行數據時發生錯誤：{str(row_error)}')
                    current_app.logger.error(f'行數據：{row.to_dict()}')
                    raise Exception(f'處理行數據時發生錯誤：{str(row_error)}')
            
            db.session.commit()
            return jsonify({'message': '成績匯入成功'})
            
        except Exception as e:
            db.session.rollback()
            raise Exception(f'讀取檔案失敗：{str(e)}')
    
    except Exception as e:
        current_app.logger.error(f'匯入成績時發生錯誤：{str(e)}')
        current_app.logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500
    
    finally:
        # 清理臨時檔案
        try:
            if temp_path and os.path.exists(temp_path):
                os.remove(temp_path)
        except Exception as e:
            current_app.logger.error(f'清理臨時檔案時發生錯誤：{str(e)}')

@bp.route('/', methods=['POST'])
def create_score():
    data = request.get_json()
    score = Score()
    score.from_dict(data)
    db.session.add(score)
    db.session.commit()
    return jsonify(score.to_dict()), 201

@bp.route('/<int:id>', methods=['PUT'])
def update_score(id):
    score = Score.query.get_or_404(id)
    data = request.get_json()
    score.from_dict(data)
    db.session.commit()
    return jsonify(score.to_dict())

@bp.route('/<int:id>', methods=['DELETE'])
def delete_score(id):
    score = Score.query.get_or_404(id)
    db.session.delete(score)
    db.session.commit()
    return '', 204

@bp.route('/clear', methods=['POST'])
def clear_scores():
    try:
        Score.query.delete()
        db.session.commit()
        return jsonify({'message': '成功清除所有成績資料'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@bp.route('/upload', methods=['POST'])
def upload_scores():
    try:
        current_app.logger.info("開始處理成績上傳請求")
        
        if 'file' not in request.files:
            current_app.logger.error("未找到上傳的文件")
            return jsonify({'error': '未找到文件'}), 400
            
        file = request.files['file']
        tournament_id = request.form.get('tournament_id')
        
        current_app.logger.info(f"接收到的文件名: {file.filename}")
        current_app.logger.info(f"賽事ID: {tournament_id}")
        
        if not tournament_id:
            current_app.logger.error("未提供賽事ID")
            return jsonify({'error': '未提供賽事ID'}), 400
            
        if file.filename == '':
            current_app.logger.error("未選擇文件")
            return jsonify({'error': '未選擇文件'}), 400
            
        if not file.filename.endswith(('.xlsx', '.xls')):
            current_app.logger.error(f"不支持的文件格式: {file.filename}")
            return jsonify({'error': '不支持的文件格式'}), 400
            
        # 保存文件到臨時目錄
        temp_path = os.path.join(current_app.instance_path, 'uploads')
        os.makedirs(temp_path, exist_ok=True)
        temp_file = os.path.join(temp_path, secure_filename(file.filename))
        file.save(temp_file)
        
        # 讀取Excel文件
        current_app.logger.info("開始讀取Excel文件")
        try:
            # 嘗試使用 openpyxl 引擎
            try:
                df = pd.read_excel(temp_file, engine='openpyxl')
            except Exception as e1:
                current_app.logger.warning(f"使用 openpyxl 引擎失敗: {str(e1)}")
                # 如果失敗，嘗試使用 xlrd 引擎
                try:
                    df = pd.read_excel(temp_file, engine='xlrd')
                except Exception as e2:
                    current_app.logger.warning(f"使用 xlrd 引擎失敗: {str(e2)}")
                    # 最後嘗試不指定引擎
                    df = pd.read_excel(temp_file)
            
            current_app.logger.info(f"Excel列名: {df.columns.tolist()}")
            
            # 清理臨時文件
            os.remove(temp_file)
            
        except Exception as e:
            current_app.logger.error(f"讀取Excel文件失敗: {str(e)}")
            if os.path.exists(temp_file):
                os.remove(temp_file)
            return jsonify({'error': '讀取Excel文件失敗', 'details': str(e)}), 400

        # 列名映射（支持多種可能的列名）
        column_mappings = {
            '會員編號': ['會員編號', '會員號碼', 'Member No', 'MemberNo'],
            'HOLE': ['HOLE', 'HOLE NAME', '全名', 'Full Name'],
            '姓名': ['姓名', '中文姓名', 'Name', 'Chinese Name'],
            '淨桿名次': ['淨桿名次', '名次', 'Rank', 'Net Rank'],
            '總桿數': ['總桿數', '總桿', 'Gross Score', 'Total'],
            '前次差點': ['前次差點', '原差點', 'Previous Handicap', 'Old Handicap'],
            '淨桿桿數': ['淨桿桿數', '淨桿', 'Net Score'],
            '差點增減': ['差點增減', '差點增减', '增減', '增减', '差點增減值', 'Handicap Change'],
            '新差點': ['新差點', '新的差點', 'New Handicap'],
            '積分': ['積分', 'Points', 'Score']
        }
        
        # 檢查每個必要的列是否存在（使用映射）
        missing_columns = []
        column_map = {}  # 存儲實際使用的列名映射
        
        for required_col, possible_names in column_mappings.items():
            found = False
            for name in possible_names:
                if name in df.columns:
                    column_map[required_col] = name
                    found = True
                    break
            if not found:
                missing_columns.append(required_col)
        
        if missing_columns:
            current_app.logger.error(f"缺少必要的列: {missing_columns}")
            return jsonify({
                'error': f'缺少必要的列: {", ".join(missing_columns)}',
                'found_columns': df.columns.tolist()  # 返回找到的列名，以便調試
            }), 400

        # 重命名列以統一處理
        df = df.rename(columns={v: k for k, v in column_map.items()})
            
        # 刪除該賽事的現有成績
        current_app.logger.info(f"刪除賽事ID {tournament_id} 的現有成績")
        Score.query.filter_by(tournament_id=tournament_id).delete()
        
        # 添加新的成績記錄
        current_app.logger.info("開始添加新的成績記錄")
        for index, row in df.iterrows():
            try:
                current_app.logger.info(f"處理第 {index + 1} 行數據")
                
                # 數據預處理和驗證
                member_number = str(row['會員編號']).strip()
                if not member_number or len(member_number) > 4:
                    raise ValueError(f"無效的會員編號: {member_number}")
                
                # 檢查會員編號格式
                if not (member_number[0].isalpha() and member_number[1:].isdigit() and len(member_number[1:]) <= 3):
                    raise ValueError(f"會員編號格式錯誤（應為一個英文字母+最多三位數字）: {member_number}")
                
                full_name = str(row['HOLE']).strip() if pd.notna(row['HOLE']) else None
                chinese_name = str(row['姓名']).strip() if pd.notna(row['姓名']) else None
                
                # 處理數值欄位
                try:
                    rank = int(row['淨桿名次']) if pd.notna(row['淨桿名次']) else None
                    gross_score = int(row['總桿數']) if pd.notna(row['總桿數']) else None
                except ValueError as e:
                    raise ValueError(f"淨桿名次或總桿數必須為整數")
                
                # 處理浮點數欄位，確保最多2位小數
                try:
                    previous_handicap = round(float(row['前次差點']), 2) if pd.notna(row['前次差點']) else None
                    net_score = round(float(row['淨桿桿數']), 2) if pd.notna(row['淨桿桿數']) else None
                    handicap_change = round(float(row['差點增減']), 2) if pd.notna(row['差點增減']) else None
                    new_handicap = round(float(row['新差點']), 2) if pd.notna(row['新差點']) else None
                except ValueError as e:
                    raise ValueError(f"差點相關欄位必須為數值")

                try:
                    points = int(row['積分']) if pd.notna(row['積分']) else None
                except ValueError as e:
                    raise ValueError(f"積分必須為整數")
                
                # 創建成績記錄
                score = Score(
                    tournament_id=tournament_id,
                    member_number=member_number,
                    full_name=full_name,
                    chinese_name=chinese_name,
                    rank=rank,
                    gross_score=gross_score,
                    previous_handicap=previous_handicap,
                    net_score=net_score,
                    handicap_change=handicap_change,
                    new_handicap=new_handicap,
                    points=points
                )
                
                db.session.add(score)
                current_app.logger.info(f"成功添加第 {index + 1} 行數據")
                
            except Exception as row_error:
                current_app.logger.error(f"處理第 {index + 1} 行數據時發生錯誤: {str(row_error)}")
                db.session.rollback()
                return jsonify({
                    'error': f'處理第 {index + 1} 行數據時發生錯誤',
                    'details': str(row_error)
                }), 400
            
        current_app.logger.info("提交所有更改到數據庫")
        db.session.commit()
        current_app.logger.info("成績上傳成功")
        return jsonify({'message': '成績上傳成功'})
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"上傳成績時發生錯誤: {str(e)}")
        current_app.logger.error(traceback.format_exc())
        return jsonify({
            'error': '上傳成績失敗',
            'details': str(e)
        }), 500

@bp.route('/annual-stats', methods=['POST'])
def get_annual_stats():
    try:
        current_app.logger.info("開始計算年度總成績")
        data = request.get_json()
        tournament_ids = data.get('tournament_ids', [])
        
        current_app.logger.info(f"接收到的賽事ID列表: {tournament_ids}")
        
        if not tournament_ids:
            return jsonify({'error': '請選擇至少一個賽事'}), 400
            
        # 獲取所選賽事的所有成績
        scores = Score.query.filter(Score.tournament_id.in_(tournament_ids)).all()
        tournaments = Tournament.query.filter(Tournament.id.in_(tournament_ids)).all()
        
        current_app.logger.info(f"找到 {len(scores)} 筆成績記錄")
        current_app.logger.info(f"找到 {len(tournaments)} 個賽事")
        
        # 建立賽事 ID 到名稱的映射
        tournament_names = {t.id: t.name for t in tournaments}
        
        # 按會員分組統計數據
        # 先按會員+賽事去重，每個會員每個賽事只保留一筆記錄
        # 為了與賽事成績頁面保持一致，我們使用ID最大的記錄（通常是最後插入的，最可能是正確的）
        # 如果同一會員同一賽事有多筆記錄，優先使用ID最大的
        unique_scores = {}
        for score in scores:
            key = (score.member_number, score.tournament_id)
            if key not in unique_scores:
                unique_scores[key] = score
            else:
                existing = unique_scores[key]
                # 保留ID最大的記錄（通常是最後插入的）
                if score.id > existing.id:
                    unique_scores[key] = score
        
        # 記錄去重後的記錄數
        current_app.logger.info(f"去重後保留 {len(unique_scores)} 筆唯一記錄")
        
        # 記錄每個會員每個賽事使用的記錄，用於調試
        for (member_number, tournament_id), score in unique_scores.items():
            current_app.logger.info(f"會員 {member_number} 賽事 {tournament_id} 使用記錄 ID {score.id}, 總桿 {score.gross_score}, 積分 {score.points}")
        
        stats = {}
        for (member_number, tournament_id), score in unique_scores.items():
            if member_number not in stats:
                stats[member_number] = {
                    'member_number': member_number,
                    'name': score.chinese_name,
                    'gender': 'F' if member_number.startswith('F') else 'M',
                    'full_name': score.full_name,
                    'total_gross_scores': [],
                    'participation_count': 0,
                    'handicaps': [],
                    'total_points': 0,
                    'tournaments': []
                }
            
            member_stats = stats[member_number]
            member_stats['participation_count'] += 1
            
            # 使用相同的記錄來計算統計資料和賽事詳情，確保一致性
            if score.gross_score is not None:
                member_stats['total_gross_scores'].append(score.gross_score)
            if score.new_handicap is not None:
                member_stats['handicaps'].append(score.new_handicap)
            if score.points is not None:
                member_stats['total_points'] += score.points
                
            # 添加個別賽事資料（使用與統計計算相同的記錄，確保一致性）
            member_stats['tournaments'].append({
                'tournament_name': tournament_names[tournament_id],
                'new_handicap': score.new_handicap,
                'gross_score': score.gross_score,  # 這應該與total_gross_scores中的值一致（當只有一個賽事時）
                'net_score': score.net_score,
                'rank': score.rank,
                'points': score.points  # 這應該與total_points一致（當只有一個賽事時）
            })
        
        # 計算平均值並格式化數據
        result = []
        for member_number, member_stats in stats.items():
            avg_gross = sum(member_stats['total_gross_scores']) / len(member_stats['total_gross_scores']) if member_stats['total_gross_scores'] else 0
            avg_handicap = sum(member_stats['handicaps']) / len(member_stats['handicaps']) if member_stats['handicaps'] else 0
            
            result.append({
                'member_number': member_number,
                'name': member_stats['name'],
                'gender': member_stats['gender'],
                'full_name': member_stats['full_name'],
                'avg_gross_score': round(avg_gross, 1),
                'participation_count': member_stats['participation_count'],
                'avg_handicap': round(avg_handicap, 1),
                'total_points': member_stats['total_points'],
                'tournaments': sorted(member_stats['tournaments'], key=lambda x: x['tournament_name'])
            })
        
        # 按總積分降序排序
        result.sort(key=lambda x: x['total_points'], reverse=True)
        
        current_app.logger.info(f"計算完成，返回 {len(result)} 筆統計數據")
        return jsonify(result)
        
    except Exception as e:
        current_app.logger.error(f"計算年度總成績時發生錯誤: {str(e)}")
        current_app.logger.error(traceback.format_exc())
        return jsonify({
            'error': '計算年度總成績失敗',
            'details': str(e)
        }), 500

@bp.route('/export/<int:tournament_id>', methods=['GET'])
def export_scores(tournament_id):
    try:
        current_app.logger.info(f"開始匯出賽事ID {tournament_id} 的成績")
        
        # 獲取賽事資訊
        tournament = Tournament.query.get_or_404(tournament_id)
        
        # 獲取該賽事的所有成績
        scores = Score.query.filter_by(tournament_id=tournament_id).all()
        
        if not scores:
            return jsonify({'error': '該賽事尚無成績記錄'}), 404
            
        # 創建一個新的工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "成績表"
        
        # 設定標題列
        headers = ['會員編號', 'HOLE', '姓名', '淨桿名次', '總桿數', 
                  '前次差點', '淨桿桿數', '差點增減', '新差點', '積分']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # 添加數據
        for row, score in enumerate(scores, 2):
            data = [
                score.member_number,
                score.full_name,
                score.chinese_name,
                score.rank,
                score.gross_score,
                score.previous_handicap,
                score.net_score,
                score.handicap_change,
                score.new_handicap,
                score.points
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = Alignment(horizontal='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # 調整欄寬
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
            
        # 將工作簿保存到 BytesIO 對象
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        # 生成檔案名稱
        filename = f"{tournament.name}_成績表_{datetime.now().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        current_app.logger.error(f"匯出成績時發生錯誤: {str(e)}")
        current_app.logger.error(traceback.format_exc())
        return jsonify({
            'error': '匯出成績失敗',
            'details': str(e)
        }), 500

@bp.route('/annual-stats/export', methods=['POST'])
def export_annual_stats():
    """匯出年度總成績為 Excel 檔案"""
    try:
        current_app.logger.info("開始匯出年度總成績")
        data = request.get_json()
        tournament_ids = data.get('tournament_ids', [])
        
        if not tournament_ids:
            return jsonify({'error': '請選擇至少一個賽事'}), 400
        
        # 獲取所選賽事的所有成績
        scores = Score.query.filter(Score.tournament_id.in_(tournament_ids)).all()
        tournaments = Tournament.query.filter(Tournament.id.in_(tournament_ids)).all()
        
        if not scores:
            return jsonify({'error': '沒有可匯出的資料'}), 404
        
        # 建立賽事 ID 到名稱的映射
        tournament_names = {t.id: t.name for t in tournaments}
        
        # 按會員分組統計數據（與 annual-stats 相同的邏輯）
        # 先按會員+賽事去重，每個會員每個賽事只保留一筆記錄
        # 為了與賽事成績頁面保持一致，使用ID最大的記錄（通常是最後插入的）
        unique_scores = {}
        for score in scores:
            key = (score.member_number, score.tournament_id)
            if key not in unique_scores:
                unique_scores[key] = score
            else:
                existing = unique_scores[key]
                # 保留ID最大的記錄（通常是最後插入的）
                if score.id > existing.id:
                    unique_scores[key] = score
        
        # 記錄去重後的記錄數
        current_app.logger.info(f"去重後保留 {len(unique_scores)} 筆唯一記錄")
        
        stats = {}
        for (member_number, tournament_id), score in unique_scores.items():
            if member_number not in stats:
                stats[member_number] = {
                    'member_number': member_number,
                    'name': score.chinese_name,
                    'gender': 'F' if member_number and member_number.startswith('F') else 'M',
                    'full_name': score.full_name,
                    'total_gross_scores': [],
                    'participation_count': 0,
                    'handicaps': [],
                    'total_points': 0,
                    'tournaments': []
                }
            
            member_stats = stats[member_number]
            member_stats['participation_count'] += 1
            
            if score.gross_score is not None:
                member_stats['total_gross_scores'].append(score.gross_score)
            if score.new_handicap is not None:
                member_stats['handicaps'].append(score.new_handicap)
            if score.points is not None:
                member_stats['total_points'] += score.points
                
            # 添加個別賽事資料
            member_stats['tournaments'].append({
                'tournament_name': tournament_names[tournament_id],
                'new_handicap': score.new_handicap,
                'gross_score': score.gross_score,
                'net_score': score.net_score,
                'rank': score.rank,
                'points': score.points
            })
        
        # 計算平均值並格式化數據
        result = []
        for member_number, member_stats in stats.items():
            avg_gross = sum(member_stats['total_gross_scores']) / len(member_stats['total_gross_scores']) if member_stats['total_gross_scores'] else 0
            avg_handicap = sum(member_stats['handicaps']) / len(member_stats['handicaps']) if member_stats['handicaps'] else 0
            
            result.append({
                'member_number': member_number,
                'name': member_stats['name'],
                'gender': member_stats['gender'],
                'full_name': member_stats['full_name'],
                'avg_gross_score': round(avg_gross, 1),
                'participation_count': member_stats['participation_count'],
                'avg_handicap': round(avg_handicap, 1),
                'total_points': member_stats['total_points'],
                'tournaments': sorted(member_stats['tournaments'], key=lambda x: x['tournament_name'])
            })
        
        # 按總積分降序排序
        result.sort(key=lambda x: x['total_points'], reverse=True)
        
        # 建立 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "年度總成績"
        
        # 設定標題行樣式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 設定標題
        headers = ['會員編號', '姓名', '性別', '級別', '總桿平均', '參與次數', '差點平均', '年度積分']
        ws.append(headers)
        
        # 設定標題行樣式
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 填入資料
        data_font = Font(size=11)
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        for row_data in result:
            row = [
                row_data['member_number'] or '',
                row_data['name'] or '',
                '女' if row_data['gender'] == 'F' else '男',
                row_data['full_name'] or '',
                row_data['avg_gross_score'],
                row_data['participation_count'],
                row_data['avg_handicap'],
                row_data['total_points']
            ]
            ws.append(row)
            
            # 設定資料行樣式
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=ws.max_row, column=col_num)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
        
        # 自動調整欄寬
        for col_num, header in enumerate(headers, 1):
            max_length = len(header)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width
        
        # 設定行高
        ws.row_dimensions[1].height = 25
        for row_num in range(2, ws.max_row + 1):
            ws.row_dimensions[row_num].height = 20
        
        # 創建第二個工作表：賽事詳情
        ws2 = wb.create_sheet(title="賽事詳情")
        
        # 設定標題
        detail_headers = ['會員編號', '姓名', '性別', '級別', '賽事名稱', '新差點', '總桿', '淨桿', '淨桿名次', '積分']
        ws2.append(detail_headers)
        
        # 設定標題行樣式
        for col_num, header in enumerate(detail_headers, 1):
            cell = ws2.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 填入賽事詳情資料（展開所有會員的賽事）
        for row_data in result:
            member_number = row_data['member_number'] or ''
            member_name = row_data['name'] or ''
            member_gender = '女' if row_data['gender'] == 'F' else '男'
            member_full_name = row_data['full_name'] or ''
            
            # 為每個會員的每個賽事添加一行
            for tournament in row_data['tournaments']:
                row = [
                    member_number,
                    member_name,
                    member_gender,
                    member_full_name,
                    tournament['tournament_name'] or '',
                    round(tournament['new_handicap'], 1) if tournament['new_handicap'] is not None else '',
                    tournament['gross_score'] if tournament['gross_score'] is not None else '',
                    round(tournament['net_score'], 1) if tournament['net_score'] is not None else '',
                    tournament['rank'] if tournament['rank'] is not None else '',
                    tournament['points'] if tournament['points'] is not None else ''
                ]
                ws2.append(row)
                
                # 設定資料行樣式
                for col_num in range(1, len(detail_headers) + 1):
                    cell = ws2.cell(row=ws2.max_row, column=col_num)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
        
        # 自動調整第二個工作表的欄寬
        for col_num, header in enumerate(detail_headers, 1):
            max_length = len(header)
            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
            adjusted_width = min(max_length + 2, 50)
            ws2.column_dimensions[get_column_letter(col_num)].width = adjusted_width
        
        # 設定第二個工作表的行高
        ws2.row_dimensions[1].height = 25
        for row_num in range(2, ws2.max_row + 1):
            ws2.row_dimensions[row_num].height = 20
        
        # 建立檔案緩衝區
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 產生檔案名稱
        tournament_names_str = '_'.join([t.name for t in sorted(tournaments, key=lambda x: x.date)][:3])
        if len(tournaments) > 3:
            tournament_names_str += f'_等{len(tournaments)}場'
        filename = f'年度總成績_{tournament_names_str}_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        current_app.logger.info(f"匯出成功，檔案名稱: {filename}")
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        current_app.logger.error(f"匯出年度總成績時發生錯誤: {str(e)}")
        current_app.logger.error(traceback.format_exc())
        return jsonify({
            'error': '匯出年度總成績失敗',
            'details': str(e)
        }), 500