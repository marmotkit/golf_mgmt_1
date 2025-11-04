from flask import Blueprint, jsonify, request, current_app, send_file
from app.models import db, TournamentAward, AwardType, Tournament
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import logging
import traceback

bp = Blueprint('awards', __name__)
logger = logging.getLogger(__name__)

@bp.before_request
def log_request_info():
    """記錄每個請求的詳細信息"""
    logger.info('====== 獎項管理 API 請求開始 ======')
    logger.info(f'請求方法: {request.method}')
    logger.info(f'請求路徑: {request.path}')
    logger.info(f'請求參數: {dict(request.args)}')
    logger.info(f'請求頭: {dict(request.headers)}')
    if request.is_json:
        logger.info(f'請求數據: {request.get_json()}')
    logger.info('================================')

@bp.route('/types', methods=['GET'])
def get_award_types():
    """獲取所有獎項類型"""
    try:
        logger.info('開始獲取獎項類型')
        award_types = AwardType.query.filter_by(is_active=True).all()
        
        # 如果沒有獎項類型，自動初始化
        if not award_types:
            logger.info('獎項類型為空，開始初始化')
            from init_award_types import init_award_types
            init_award_types()
            award_types = AwardType.query.filter_by(is_active=True).all()
            
        logger.info(f'找到 {len(award_types)} 個獎項類型')
        result = [t.to_dict() for t in award_types]
        logger.info(f'返回數據: {result}')
        return jsonify(result)
    except Exception as e:
        logger.error(f"獲取獎項類型時發生錯誤: {str(e)}")
        logger.error(f"錯誤類型: {type(e).__name__}")
        logger.error(f"完整錯誤信息: {traceback.format_exc()}")
        return jsonify({
            'error': '獲取獎項類型失敗',
            'message': str(e),
            'traceback': traceback.format_exc()
        }), 500

@bp.route('/types', methods=['POST'])
def create_award_type():
    """創建新的獎項類型"""
    try:
        data = request.get_json()
        award_type = AwardType(**data)
        db.session.add(award_type)
        db.session.commit()
        return jsonify(award_type.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        logger.error(f"創建獎項類型時發生錯誤: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@bp.route('/types/<int:id>', methods=['PUT'])
def update_award_type(id):
    """更新獎項類型"""
    try:
        award_type = AwardType.query.get_or_404(id)
        data = request.get_json()
        for key, value in data.items():
            setattr(award_type, key, value)
        db.session.commit()
        return jsonify(award_type.to_dict())
    except Exception as e:
        db.session.rollback()
        logger.error(f"更新獎項類型時發生錯誤: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@bp.route('/', methods=['GET'])
def get_tournament_awards():
    """獲取賽事獎項"""
    try:
        logger.info('開始獲取賽事獎項')
        tournament_id = request.args.get('tournament_id', type=int)
        logger.info(f'賽事ID: {tournament_id}')
        
        if tournament_id:
            # 如果提供了賽事ID，只返回該賽事的獎項
            logger.info(f'查詢賽事 {tournament_id} 的獎項')
            awards = TournamentAward.query.filter_by(tournament_id=tournament_id).all()
        else:
            # 如果沒有提供賽事ID，返回所有獎項
            logger.info('查詢所有賽事的獎項')
            awards = TournamentAward.query.all()
            
        logger.info(f'找到 {len(awards)} 個獎項')
        
        result = []
        for award in awards:
            try:
                award_dict = award.to_dict()
                result.append(award_dict)
            except Exception as e:
                logger.error(f'轉換獎項 {award.id} 為字典時發生錯誤: {str(e)}')
                logger.error(traceback.format_exc())
                continue
        
        logger.info(f'返回數據: {result}')
        return jsonify(result)
    except Exception as e:
        logger.error(f"獲取賽事獎項時發生錯誤: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'error': '獲取賽事獎項失敗',
            'message': str(e),
            'traceback': traceback.format_exc()
        }), 500

@bp.route('/', methods=['POST'])
def create_tournament_award():
    """創建賽事獎項"""
    try:
        data = request.get_json()
        award = TournamentAward(**data)
        db.session.add(award)
        db.session.commit()
        return jsonify(award.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        logger.error(f"創建賽事獎項時發生錯誤: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@bp.route('/<int:id>', methods=['PUT'])
def update_tournament_award(id):
    """更新賽事獎項"""
    try:
        award = TournamentAward.query.get_or_404(id)
        data = request.get_json()
        for key, value in data.items():
            setattr(award, key, value)
        db.session.commit()
        return jsonify(award.to_dict())
    except Exception as e:
        db.session.rollback()
        logger.error(f"更新賽事獎項時發生錯誤: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@bp.route('/<int:id>', methods=['DELETE'])
def delete_tournament_award(id):
    """刪除賽事獎項"""
    try:
        award = TournamentAward.query.get_or_404(id)
        db.session.delete(award)
        db.session.commit()
        return '', 204
    except Exception as e:
        db.session.rollback()
        logger.error(f"刪除賽事獎項時發生錯誤: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@bp.route('/batch', methods=['POST'])
def batch_create_awards():
    """批量創建賽事獎項"""
    try:
        data = request.get_json()
        awards = []
        for award_data in data:
            award = TournamentAward(**award_data)
            db.session.add(award)
            awards.append(award)
        db.session.commit()
        return jsonify([a.to_dict() for a in awards]), 201
    except Exception as e:
        db.session.rollback()
        logger.error(f"批量創建賽事獎項時發生錯誤: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@bp.route('/export', methods=['GET'])
def export_awards():
    """匯出獎項資料為 Excel 檔案"""
    try:
        tournament_name = request.args.get('tournament_name', '')
        
        if not tournament_name:
            return jsonify({'error': '請指定賽事名稱'}), 400
        
        logger.info(f'開始匯出賽事 "{tournament_name}" 的獎項資料')
        
        # 根據賽事名稱找到賽事
        tournament = Tournament.query.filter_by(name=tournament_name).first()
        if not tournament:
            return jsonify({'error': '找不到指定的賽事'}), 404
        
        # 獲取該賽事的所有獎項
        awards = TournamentAward.query.filter_by(tournament_id=tournament.id).all()
        
        if not awards:
            return jsonify({'error': '沒有可匯出的資料'}), 404
        
        # 建立 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "獎項管理"
        
        # 設定標題行樣式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 設定標題
        headers = ['獎項類型', '得獎者姓名', '排名', '備註']
        ws.append(headers)
        
        # 設定標題行樣式
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 填入資料
        data_font = Font(size=11)
        data_alignment = Alignment(horizontal="left", vertical="center")
        
        for award in awards:
            # 獲取獎項類型名稱
            award_type = AwardType.query.get(award.award_type_id)
            award_type_name = award_type.name if award_type else '未知'
            
            row = [
                award_type_name,
                award.chinese_name or '',
                award.rank if award.rank else '',
                award.remarks or ''
            ]
            ws.append(row)
            
            # 設定資料行樣式
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=ws.max_row, column=col_num)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
        
        # 自動調整欄寬
        for col_num, header in enumerate(headers, 1):
            max_length = len(header)
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_num, max_col=col_num):
                for cell in row:
                    if cell.value:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[get_column_letter(col_num)].width = adjusted_width
        
        # 設定行高
        ws.row_dimensions[1].height = 25
        for row_num in range(2, ws.max_row + 1):
            ws.row_dimensions[row_num].height = 20
        
        # 建立檔案緩衝區
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 產生檔案名稱
        filename = f'獎項管理_{tournament_name}_{datetime.now().strftime("%Y%m%d")}.xlsx'
        
        logger.info(f'匯出成功，檔案名稱: {filename}')
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"匯出獎項資料失敗: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({
            'error': '匯出獎項資料失敗',
            'details': str(e)
        }), 500 