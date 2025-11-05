from flask import Blueprint, jsonify, request, current_app, send_file
from werkzeug.utils import secure_filename
import pandas as pd
import os
from app import db
from app.models import Member, MemberVersion
import logging
import traceback
from datetime import datetime
from sqlalchemy import func
from sqlalchemy import Boolean
import io
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

bp = Blueprint('members', __name__)

# Configure logging
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

def ensure_upload_dir():
    upload_dir = current_app.config['UPLOAD_FOLDER']
    if not os.path.exists(upload_dir):
        os.makedirs(upload_dir)
    return upload_dir

def generate_version_number():
    """生成新的版本號"""
    try:
        # 獲取最新版本號
        latest_version = db.session.query(MemberVersion.version)\
            .order_by(MemberVersion.version.desc())\
            .first()
        
        if latest_version:
            current_version = latest_version[0]
        else:
            # 如果沒有任何版本，從202501001開始
            current_date = datetime.now().strftime('%Y%m%d')
            current_version = f"{current_date}001"
            
        # 生成新版本號
        date_part = str(current_version)[:8]  # 取出日期部分
        current_date = datetime.now().strftime('%Y%m%d')
        
        if date_part == current_date:
            # 如果是同一天，序號加1
            sequence = int(str(current_version)[8:]) + 1
        else:
            # 如果是新的一天，序號重置為1
            sequence = 1
            
        new_version = f"{current_date}{sequence:03d}"  # 確保序號是3位數
        logger.info(f'Generated new version number: {new_version} (current: {current_version})')
        return new_version
        
    except Exception as e:
        logger.error(f'Error generating version number: {str(e)}')
        logger.error(traceback.format_exc())
        raise

def process_excel_data(df):
    """處理上傳的 Excel 資料"""
    try:
        # 檢查必要欄位是否存在
        required_columns = {
            '會員編號': ['會員編號', '會員號碼'],
            '會員類型': ['會員類型', '類型', '會員/來賓'],  # 添加新的欄位名稱
            '帳號': ['帳號', '帳戶', 'account'],
            '是否為管理員': ['是否為管理員', '管理員'],
            '性別': ['性別', 'gender'],
            '差點': ['差點', '最新差點']  # 添加新的欄位名稱
        }

        # 檢查每個必要欄位
        column_mapping = {}
        missing_columns = []
        available_columns = df.columns.tolist()
        
        logger.info(f'Excel檔案中的欄位：{available_columns}')
        
        for required_col, possible_names in required_columns.items():
            found = False
            for name in possible_names:
                if name in df.columns:
                    column_mapping[required_col] = name
                    found = True
                    break
            if not found:
                missing_columns.append(required_col)

        if missing_columns:
            error_msg = f'缺少必要欄位：{", ".join(missing_columns)}\n可用的欄位：{", ".join(available_columns)}'
            logger.error(error_msg)
            raise ValueError(error_msg)

        # 重命名欄位以統一處理
        df = df.rename(columns={v: k for k, v in column_mapping.items()})

        # 清理數據：移除前後空格，將 '(null)' 轉換為空字串
        for col in df.columns:
            if df[col].dtype == object:
                df[col] = df[col].astype(str).str.strip()
                df[col] = df[col].replace(['(null)', 'null', 'nan', 'None', ''], '')

        # 驗證每一行數據
        errors = []
        valid_rows = []
        
        for idx, row in df.iterrows():
            try:
                row_errors = []
                row_data = row.to_dict()
                
                # 檢查必填欄位是否為空
                empty_fields = []
                for col in required_columns.keys():
                    if pd.isna(row[col]) or str(row[col]).strip() == '':
                        empty_fields.append(col)
                
                if empty_fields:
                    row_errors.append(f'第 {idx + 2} 行：必填欄位為空：{", ".join(empty_fields)}')

                # 驗證會員編號格式
                member_number = str(row['會員編號']).strip()
                if len(member_number) == 0:
                    row_errors.append(f'第 {idx + 2} 行：會員編號不能為空')
                elif len(member_number) > 10:
                    row_errors.append(f'第 {idx + 2} 行：會員編號長度不能超過10個字符：{member_number}')

                # 驗證會員類型
                member_type = str(row['會員類型']).strip()
                if member_type and member_type not in ['會員', '來賓']:
                    row_errors.append(f'第 {idx + 2} 行：會員類型必須是「會員」或「來賓」，目前值為：{member_type}')

                # 驗證管理員欄位
                is_admin = str(row['是否為管理員']).strip()
                if is_admin and is_admin not in ['是', '否']:
                    row_errors.append(f'第 {idx + 2} 行：是否為管理員必須是「是」或「否」，目前值為：{is_admin}')

                # 驗證性別
                gender = str(row['性別']).strip()
                if gender and gender not in ['M', 'F']:
                    row_errors.append(f'第 {idx + 2} 行：性別必須是「M」或「F」，目前值為：{gender}')

                # 驗證差點格式（如果存在）
                if '差點' in row and str(row['差點']).strip():
                    try:
                        handicap = float(str(row['差點']).strip())
                    except ValueError:
                        row_errors.append(f'第 {idx + 2} 行：差點必須是數字，目前值為：{row["差點"]}')

                # 如果有錯誤，加入到錯誤列表
                if row_errors:
                    errors.extend(row_errors)
                else:
                    valid_rows.append(row)
                
            except Exception as e:
                logger.error(f'處理第 {idx + 2} 行時發生異常：{str(e)}')
                logger.error(f'該行資料內容：{row.to_dict()}')
                errors.append(f'第 {idx + 2} 行處理失敗：{str(e)}')

        if errors:
            error_summary = '\n'.join(errors)
            logger.error(f'資料驗證發現錯誤：\n{error_summary}')
            raise ValueError(error_summary)

        # 創建新的 DataFrame，只包含有效的資料行
        valid_df = pd.DataFrame(valid_rows)
        logger.info(f'成功處理 {len(valid_rows)} 行資料')
        return valid_df

    except Exception as e:
        logger.error(f'處理 Excel 數據時發生錯誤：{str(e)}')
        raise ValueError(f'處理 Excel 數據時發生錯誤：{str(e)}')

def _parse_bool(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.lower() in ('true', '1', 'yes', 'on')
    if isinstance(value, int):
        return bool(value)
    return False

@bp.route('/', methods=['GET'])
def get_members():
    """獲取會員列表"""
    try:
        # 從查詢參數中獲取版本號，如果沒有指定則使用最新版本
        version = request.args.get('version')
        
        # 獲取最新版本號
        latest_version = db.session.query(MemberVersion.version)\
            .order_by(MemberVersion.version.desc())\
            .first()
            
        if not version and latest_version:
            version = latest_version[0]
        elif not latest_version:
            logger.warning('No member versions found in database')
            return jsonify([])
            
        logger.info(f'Fetching members for version: {version}')
        
        # 使用指定版本號獲取會員資料
        members_data = db.session.query(
            Member,
            MemberVersion.data
        ).join(
            MemberVersion,
            Member.id == MemberVersion.member_id
        ).filter(
            MemberVersion.version == version
        ).all()

        # 整理會員資料
        result = []
        for member, version_data in members_data:
            try:
                member_dict = {
                    'id': member.id,
                    'account': member.account,
                    'chinese_name': member.chinese_name,
                    'english_name': member.english_name,
                    'department_class': member.department_class,
                    'member_number': member.member_number,
                    'is_guest': member.is_guest,
                    'is_admin': member.is_admin,
                    'gender': member.gender,
                    'handicap': float(version_data.get('handicap')) if version_data and version_data.get('handicap') is not None else None
                }
                result.append(member_dict)
            except Exception as e:
                logger.error(f'Error processing member {member.id}: {str(e)}')
                continue

        logger.info(f'Found {len(result)} members for version {version}')
        return jsonify(result)

    except Exception as e:
        logger.error(f'Error getting members: {str(e)}')
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@bp.route('/upload', methods=['POST'])
def upload_members():
    logger.info('File upload request received')
    logger.info(f'Request files: {request.files}')
    logger.info(f'Request form: {request.form}')
    
    try:
        if 'file' not in request.files:
            error_msg = '未提供檔案'
            logger.error(error_msg)
            return jsonify({
                'error': error_msg,
                'error_messages': ['請選擇要上傳的 Excel 檔案'],
                'success_count': 0
            }), 400
            
        file = request.files['file']
        logger.info(f'Received file: {file.filename}')
        
        if file.filename == '':
            error_msg = '未選擇檔案'
            logger.error(error_msg)
            return jsonify({
                'error': error_msg,
                'error_messages': ['請選擇要上傳的 Excel 檔案'],
                'success_count': 0
            }), 400
            
        if not file.filename.endswith(('.xlsx', '.xls')):
            error_msg = '檔案格式錯誤'
            logger.error(f'{error_msg}: {file.filename}')
            return jsonify({
                'error': error_msg,
                'error_messages': ['只允許上傳 .xlsx 或 .xls 格式的檔案'],
                'success_count': 0
            }), 400

        # 確保上傳目錄存在
        upload_dir = ensure_upload_dir()
        logger.info(f'Upload directory: {upload_dir}')
        
        # 使用安全的檔案名稱儲存檔案
        filename = secure_filename(file.filename)
        filepath = os.path.join(upload_dir, filename)
        file.save(filepath)
        logger.info(f'File saved to: {filepath}')
        
        try:
            logger.info('Reading Excel file...')
            # 讀取 Excel 檔案時不過濾空值
            df = pd.read_excel(filepath, dtype=str, na_filter=False)
            logger.info(f'Excel file read successfully: {len(df)} rows')
            logger.info(f'Columns found in file: {df.columns.tolist()}')
            
            if len(df) == 0:
                error_msg = 'Excel 檔案為空'
                logger.error(error_msg)
                return jsonify({
                    'error': error_msg,
                    'error_messages': ['上傳的 Excel 檔案沒有任何資料'],
                    'success_count': 0
                }), 400
                
            # 檢查必要欄位
            required_columns = {
                '會員編號': ['會員編號', '會員號碼', 'Member ID', 'MemberID'],
                '會員類型': ['會員類型', '類型', '會員/來賓', 'Member Type'],
                '帳號': ['帳號', '帳戶', 'account', 'Account'],
                '是否為管理員': ['是否為管理員', '管理員', 'Is Administrator?', 'Is Admin'],
                '性別': ['性別', 'gender', 'Gender'],
                '差點': ['差點', '最新差點', 'Handicap', 'Latest Handicap']
            }
            
            # 可選欄位的映射（支持多種可能的欄位名稱）
            optional_columns = {
                '中文姓名': ['中文姓名', 'Chinese Name', '中文名'],
                '英文姓名': ['英文姓名', 'English Name', '英文名'],
                '系級': ['系級', 'Department/Grade', '系所', 'Department']
            }

            # 檢查每個必要欄位
            missing_columns = []
            column_mapping = {}
            available_columns = df.columns.tolist()
            
            logger.info(f'Available columns in file: {available_columns}')
            
            for required_col, possible_names in required_columns.items():
                found = False
                for name in possible_names:
                    if name in df.columns:
                        column_mapping[required_col] = name
                        found = True
                        break
                if not found:
                    missing_columns.append(required_col)

            if missing_columns:
                error_msg = f'缺少必要欄位：{", ".join(missing_columns)}\n可用的欄位：{", ".join(available_columns)}'
                logger.error(error_msg)
                return jsonify({
                    'error': '缺少必要欄位',
                    'error_messages': [error_msg],
                    'success_count': 0
                }), 400
            
            # 處理可選欄位：在重命名必要欄位之前尋找並建立映射
            # 記錄哪些欄位名稱已經被必要欄位使用
            used_column_names = set(column_mapping.values())
            
            optional_mapping = {}
            for standard_col, possible_names in optional_columns.items():
                # 如果標準名稱已經被必要欄位使用，跳過
                if standard_col in column_mapping:
                    continue
                    
                # 尋找可選欄位的原始名稱
                for name in possible_names:
                    # 如果這個名稱存在於 DataFrame 且沒有被必要欄位使用
                    if name in df.columns and name not in used_column_names:
                        optional_mapping[standard_col] = name
                        break
                
            # 合併所有映射，一次性重命名所有欄位
            all_column_mapping = {**column_mapping, **optional_mapping}
            # 建立反向映射：原始名稱 -> 標準名稱
            rename_dict = {v: k for k, v in all_column_mapping.items()}
            df = df.rename(columns=rename_dict)

            # 清理和驗證數據
            error_messages = []
            valid_rows = []
            
            for idx, row in df.iterrows():
                row_errors = []
                
                # 檢查必填欄位
                for col in required_columns.keys():
                    if pd.isna(row[col]) or str(row[col]).strip() == '':
                        row_errors.append(f'第 {idx + 2} 行：{col} 不能為空')

                # 驗證會員類型
                member_type = str(row['會員類型']).strip()
                if member_type and member_type not in ['會員', '來賓']:
                    row_errors.append(f'第 {idx + 2} 行：會員類型必須是「會員」或「來賓」，目前值為：{member_type}')

                # 驗證管理員欄位
                is_admin = str(row['是否為管理員']).strip()
                if is_admin and is_admin not in ['是', '否']:
                    row_errors.append(f'第 {idx + 2} 行：是否為管理員必須是「是」或「否」，目前值為：{is_admin}')

                # 驗證性別
                gender = str(row['性別']).strip()
                if gender and gender not in ['M', 'F']:
                    row_errors.append(f'第 {idx + 2} 行：性別必須是「M」或「F」，目前值為：{gender}')

                # 驗證差點格式（如果存在）
                if '差點' in row and str(row['差點']).strip():
                    try:
                        handicap = float(str(row['差點']).strip())
                    except ValueError:
                        row_errors.append(f'第 {idx + 2} 行：差點必須是數字，目前值為：{row["差點"]}')

                if row_errors:
                    error_messages.extend(row_errors)
                else:
                    valid_rows.append(row)

            if error_messages:
                logger.error('Validation errors found')
                logger.error('\n'.join(error_messages))
                return jsonify({
                    'error': '資料驗證失敗',
                    'error_messages': error_messages,
                    'success_count': 0
                }), 400
            
            # 處理有效的資料
            version_number = generate_version_number()
            
            # 不再刪除所有資料，而是保留舊版本
            success_count = 0
            
            # 建立會員資料字典，用於快速查找（按會員編號和帳號）
            all_members = Member.query.all()
            existing_by_number = {
                member.member_number: member 
                for member in all_members
                if member.member_number
            }
            existing_by_account = {
                member.account: member 
                for member in all_members
                if member.account
            }
            
            # 處理每一行資料
            processing_errors = []  # 累積處理錯誤
            
            for idx, row in enumerate(valid_rows):
                try:
                    member_number = str(row['會員編號']).strip()
                    account = str(row['帳號']).strip()
                    
                    # 檢查會員是否已存在（按會員編號或帳號）
                    member = existing_by_number.get(member_number)
                    
                    # 如果按會員編號找不到，但帳號存在，檢查是否應該更新
                    if not member and account:
                        member = existing_by_account.get(account)
                        if member:
                            # 檢查現有會員的會員編號
                            existing_member_number = str(member.member_number) if member.member_number else None
                            # 如果找到的會員編號不同或為空，更新會員編號
                            if existing_member_number != member_number:
                                # 檢查新的會員編號是否已被其他會員使用
                                if member_number in existing_by_number:
                                    error_msg = f'第 {idx + 2} 行：會員編號 {member_number} 已被其他會員使用，但帳號 {account} 已存在於會員編號 {existing_member_number}'
                                    logger.warning(error_msg)
                                    processing_errors.append(error_msg)
                                    continue
                                # 更新會員編號
                                old_number = existing_member_number or '(無)'
                                member.member_number = member_number
                                existing_by_number[member_number] = member
                                # 如果舊的會員編號存在，從字典中移除
                                if existing_member_number and existing_member_number in existing_by_number:
                                    del existing_by_number[existing_member_number]
                                logger.info(f'更新會員 {account} 的會員編號從 {old_number} 到 {member_number}')
                    
                    if not member:
                        # 如果會員不存在，創建新會員
                        # 安全地獲取可選欄位，如果欄位不存在則返回空字串
                        try:
                            chinese_name = str(row['中文姓名']).strip() if '中文姓名' in row and pd.notna(row['中文姓名']) else ''
                        except (KeyError, TypeError):
                            chinese_name = ''
                        
                        try:
                            english_name = str(row['英文姓名']).strip() if '英文姓名' in row and pd.notna(row['英文姓名']) else ''
                        except (KeyError, TypeError):
                            english_name = ''
                        
                        try:
                            department_class = str(row['系級']).strip() if '系級' in row and pd.notna(row['系級']) else ''
                        except (KeyError, TypeError):
                            department_class = ''
                        
                        # 檢查帳號是否已被使用（在創建新會員前）
                        if account and account in existing_by_account:
                            error_msg = f'第 {idx + 2} 行：帳號 {account} 已被會員編號 {existing_by_account[account].member_number} 使用，無法創建新會員 {member_number}'
                            logger.error(error_msg)
                            processing_errors.append(error_msg)
                            continue
                        
                        member = Member(
                            member_number=member_number,
                            account=account,
                            chinese_name=chinese_name,
                            english_name=english_name,
                            department_class=department_class,
                            is_guest=str(row['會員類型']).strip() == '來賓',
                            is_admin=str(row['是否為管理員']).strip() == '是',
                            gender=str(row['性別']).strip()
                        )
                        db.session.add(member)
                        db.session.flush()  # 獲取新會員的 ID
                        # 更新現有會員字典，以便後續查找
                        existing_by_number[member_number] = member
                        if account:
                            existing_by_account[account] = member
                    else:
                        # 如果會員已存在，更新會員資料
                        try:
                            if '中文姓名' in row and pd.notna(row['中文姓名']) and str(row['中文姓名']).strip():
                                member.chinese_name = str(row['中文姓名']).strip()
                        except (KeyError, TypeError):
                            pass
                        
                        try:
                            if '英文姓名' in row and pd.notna(row['英文姓名']) and str(row['英文姓名']).strip():
                                member.english_name = str(row['英文姓名']).strip()
                        except (KeyError, TypeError):
                            pass
                        
                        try:
                            if '系級' in row and pd.notna(row['系級']) and str(row['系級']).strip():
                                member.department_class = str(row['系級']).strip()
                        except (KeyError, TypeError):
                            pass
                        
                        # 更新其他可能變更的欄位
                        # 注意：如果帳號改變，需要更新 existing_by_account 字典
                        old_account = member.account
                        new_account = str(row['帳號']).strip()
                        
                        if new_account and new_account != old_account:
                            # 檢查新帳號是否已被使用
                            if new_account in existing_by_account and existing_by_account[new_account].id != member.id:
                                error_msg = f'第 {idx + 2} 行：帳號 {new_account} 已被其他會員使用'
                                logger.error(error_msg)
                                processing_errors.append(error_msg)
                                continue
                            # 更新帳號
                            member.account = new_account
                            # 更新字典
                            if old_account and old_account in existing_by_account:
                                del existing_by_account[old_account]
                            if new_account:
                                existing_by_account[new_account] = member
                        
                        member.is_guest = str(row['會員類型']).strip() == '來賓'
                        member.is_admin = str(row['是否為管理員']).strip() == '是'
                        member.gender = str(row['性別']).strip()
                    
                    # 創建新的版本記錄
                    handicap_value = None
                    if '差點' in row and pd.notna(row['差點']) and str(row['差點']).strip():
                        try:
                            handicap_value = float(str(row['差點']).strip())
                        except (ValueError, TypeError):
                            logger.warning(f'第 {idx + 2} 行：無法解析差點值：{row["差點"]}')
                    
                    version_data = {
                        'handicap': handicap_value
                    }
                    
                    member_version = MemberVersion(
                        member_id=member.id,
                        version=version_number,
                        data=version_data
                    )
                    db.session.add(member_version)
                    success_count += 1
                    
                except KeyError as e:
                    error_msg = f'第 {idx + 2} 行：缺少必要欄位 {str(e)}'
                    logger.error(error_msg)
                    try:
                        logger.error(f'該行可用欄位：{list(row.keys())}')
                    except:
                        pass
                    processing_errors.append(error_msg)
                    # 繼續處理下一行，但記錄錯誤
                    continue
                except Exception as e:
                    error_msg = f'第 {idx + 2} 行處理失敗：{str(e)}'
                    logger.error(error_msg)
                    try:
                        row_dict = row.to_dict() if hasattr(row, "to_dict") else dict(row)
                        logger.error(f'該行資料內容：{row_dict}')
                    except:
                        pass
                    logger.error(traceback.format_exc())
                    processing_errors.append(error_msg)
                    # 繼續處理下一行，但記錄錯誤
                    continue
            
            # 如果有處理錯誤且沒有成功處理任何記錄，返回錯誤
            if processing_errors and success_count == 0:
                logger.error(f'所有行處理都失敗，共 {len(processing_errors)} 個錯誤')
                return jsonify({
                    'error': '所有資料處理都失敗',
                    'error_messages': processing_errors[:10],  # 最多返回前10個錯誤
                    'success_count': 0
                }), 400
            
            try:
                db.session.commit()
                logger.info(f'Successfully processed {success_count} members')
                return jsonify({
                    'message': '上傳成功',
                    'version': version_number,
                    'success_count': success_count
                })
            except Exception as e:
                db.session.rollback()
                logger.error(f'Error committing changes: {str(e)}')
                return jsonify({
                    'error': '儲存資料時發生錯誤',
                    'error_messages': [str(e)],
                    'success_count': 0
                }), 500

        except pd.errors.EmptyDataError:
            error_msg = 'Excel 檔案為空'
            logger.error(error_msg)
            return jsonify({
                'error': error_msg,
                'error_messages': ['上傳的 Excel 檔案沒有任何資料'],
                'success_count': 0
            }), 400
        except Exception as e:
            error_msg = f'資料處理過程中發生錯誤：{str(e)}'
            logger.error(error_msg)
            logger.error(traceback.format_exc())
            return jsonify({
                'error': error_msg,
                'error_messages': [str(e)],
                'success_count': 0
            }), 400
            
    except Exception as e:
        error_msg = f'系統錯誤：{str(e)}'
        logger.error(error_msg)
        logger.error(traceback.format_exc())
        return jsonify({
            'error': error_msg,
            'error_messages': [str(e)],
            'success_count': 0
        }), 500

    finally:
        # 清理臨時檔案
        try:
            if 'filepath' in locals() and os.path.exists(filepath):
                os.remove(filepath)
                logger.info(f'Temporary file removed: {filepath}')
        except Exception as e:
            logger.error(f'Error removing temporary file: {str(e)}')

@bp.route('/batch-delete', methods=['POST'])
def batch_delete_members():
    logger.info('Batch delete request received')
    data = request.get_json()
    if not data or 'ids' not in data:
        logger.error('No member IDs provided')
        return jsonify({'error': 'No member IDs provided'}), 400
        
    member_ids = data['ids']
    deleted_count = 0
    error_ids = []
    
    for member_id in member_ids:
        try:
            member = Member.query.get(member_id)
            if member:
                db.session.delete(member)
                deleted_count += 1
        except Exception as e:
            logger.error(f'Error deleting member {member_id}: {str(e)}')
            logger.error(traceback.format_exc())
            error_ids.append({
                'id': member_id,
                'error': str(e)
            })
    
    db.session.commit()
    logger.info(f'Deletion completed: {deleted_count} members deleted successfully')
    
    return jsonify({
        'deleted_count': deleted_count,
        'error_count': len(error_ids),
        'errors': error_ids
    })

@bp.route('/', methods=['POST'])
def create_member():
    logger.info('Create member request received')
    data = request.get_json()
    try:
        member = Member(
            account=data['account'],
            chinese_name=data['chinese_name'],
            english_name=data.get('english_name'),
            department_class=data.get('department_class'),
            member_number=data.get('member_number'),
            is_guest=data.get('is_guest', False),
            is_admin=data.get('is_admin', False),
            gender=data.get('gender'),
            handicap=data.get('handicap')
        )
        db.session.add(member)
        
        # Create initial version
        version = MemberVersion(
            member=member,
            version=generate_version_number(),
            data=data,
            created_at=datetime.now()
        )
        db.session.add(version)
        
        db.session.commit()
        logger.info(f'Member created: {member.id}')
        return jsonify({'id': member.id}), 201
    except Exception as e:
        logger.error(f'Error creating member: {str(e)}')
        logger.error(traceback.format_exc())
        return jsonify({'error': f'Error creating member: {str(e)}'}), 400

@bp.route('/<int:member_id>', methods=['PUT', 'PATCH'])
def update_member(member_id):
    try:
        data = request.get_json()
        logger.info(f'Received update request for member {member_id}')
        logger.info(f'Request data: {data}')
        
        if data is None:
            logger.error('No JSON data received')
            return jsonify({'error': '無效的請求數據'}), 400
            
        member = Member.query.get(member_id)
        if not member:
            logger.error(f'Member {member_id} not found')
            return jsonify({'error': f'找不到會員 ID {member_id}'}), 404
            
        logger.info(f'Found member: {member.member_number}')
        logger.info(f'Current member data: {member.to_dict()}')
        
        # 移除 id 欄位，避免嘗試更新主鍵
        if 'id' in data:
            del data['id']
            
        # 更新會員資料
        if 'member_number' in data:
            if str(data['member_number']) != str(member.member_number):
                existing = Member.query.filter_by(member_number=str(data['member_number'])).first()
                if existing and existing.id != member_id:
                    return jsonify({'error': f"會員編號 {data['member_number']} 已存在"}), 400
            member.member_number = str(data['member_number'])
            
        if 'account' in data:
            member.account = str(data['account'])
            
        if 'chinese_name' in data:
            member.chinese_name = str(data['chinese_name'])
            
        if 'english_name' in data:
            member.english_name = str(data['english_name'])
            
        if 'department_class' in data:
            member.department_class = str(data['department_class'])
            
        if 'is_guest' in data:
            member.is_guest = bool(data['is_guest'])
            
        if 'is_admin' in data:
            member.is_admin = bool(data['is_admin'])
            
        if 'gender' in data:
            member.gender = str(data['gender'])
            
        if 'handicap' in data:
            try:
                handicap_value = data['handicap']
                logger.info(f'Processing handicap value: {handicap_value} (type: {type(handicap_value)})')
                if handicap_value is None or handicap_value == '':
                    member.handicap = None
                    logger.info('Setting handicap to None')
                else:
                    try:
                        member.handicap = float(handicap_value)
                        logger.info(f'Setting handicap to {member.handicap}')
                    except (ValueError, TypeError) as e:
                        logger.error(f'Failed to convert handicap to float: {str(e)}')
                        return jsonify({'error': f'差點必須是數字，收到的值為: {handicap_value}'}), 400
            except Exception as e:
                logger.error(f'Error processing handicap: {str(e)}')
                logger.error(traceback.format_exc())
                return jsonify({'error': f'處理差點時發生錯誤: {str(e)}'}), 400
        
        # 獲取最新版本
        latest_version = db.session.query(MemberVersion)\
            .filter_by(member_id=member_id)\
            .order_by(MemberVersion.version.desc())\
            .first()
        
        if not latest_version:
            logger.error(f'No version found for member {member_id}')
            return jsonify({'error': f'找不到會員 {member_id} 的版本資料'}), 404
        
        # 更新當前版本的資料
        member_data = member.to_dict()
        version_data = {k: v for k, v in member_data.items() if k not in ['created_at', 'updated_at']}
        latest_version.data = version_data
        
        logger.info('Committing changes to database')
        db.session.commit()
        logger.info(f'Successfully updated member {member_id}')
        return jsonify({
            'message': '更新成功',
            'data': version_data
        })
            
    except Exception as e:
        logger.error(f"更新會員失敗: {str(e)}")
        logger.error(traceback.format_exc())
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@bp.route('/<int:id>', methods=['DELETE'])
def delete_member(id):
    try:
        member = Member.query.get_or_404(id)
        db.session.delete(member)
        db.session.commit()
        return jsonify({'message': '刪除成功'})
    except Exception as e:
        db.session.rollback()
        logger.error(f"刪除會員失敗: {str(e)}")
        return jsonify({'error': str(e)}), 400

@bp.route('/versions/<int:id>', methods=['GET'])
def get_member_versions(id):
    logger.info(f'Get member versions request received for member {id}')
    versions = MemberVersion.query.filter_by(member_id=id).order_by(
        MemberVersion.version.desc()).all()
    return jsonify([{
        'version': v.version,
        'data': v.data,
        'created_at': v.created_at
    } for v in versions])

@bp.route('/compare', methods=['POST'])
def compare_versions():
    """比較兩個版本的差異"""
    try:
        data = request.get_json()
        from_version = data.get('from')  # 舊版本
        to_version = data.get('to')      # 新版本

        if not from_version or not to_version:
            return jsonify({'error': '請提供要比較的版本號'}), 400

        # 獲取兩個版本的會員資料
        old_data = db.session.query(
            Member.member_number,
            Member.chinese_name,
            MemberVersion.data
        ).join(
            MemberVersion,
            Member.id == MemberVersion.member_id
        ).filter(
            MemberVersion.version == from_version
        ).all()

        new_data = db.session.query(
            Member.member_number,
            Member.chinese_name,
            MemberVersion.data
        ).join(
            MemberVersion,
            Member.id == MemberVersion.member_id
        ).filter(
            MemberVersion.version == to_version
        ).all()

        # 將資料轉換為字典格式，方便比較
        old_dict = {m.member_number: m.data for m in old_data}
        new_dict = {m.member_number: m.data for m in new_data}

        # 比較差異
        differences = []
        all_members = set(old_dict.keys()) | set(new_dict.keys())

        for member_number in all_members:
            old_member_data = old_dict.get(member_number, {})
            new_member_data = new_dict.get(member_number, {})

            # 比較所有欄位
            for field in ['handicap', 'is_guest', 'is_admin', 'chinese_name', 'english_name', 'department_class', 'gender']:
                old_value = old_member_data.get(field)
                new_value = new_member_data.get(field)

                if old_value != new_value:
                    # 找到對應的會員姓名
                    member_name = next(
                        (m.chinese_name for m in old_data if m.member_number == member_number),
                        next(
                            (m.chinese_name for m in new_data if m.member_number == member_number),
                            None
                        )
                    )

                    differences.append({
                        'member_number': member_number,
                        'name': member_name,
                        'field': field,
                        'old': old_value,
                        'new': new_value
                    })

        logger.info(f'Found {len(differences)} differences between versions {from_version} and {to_version}')
        return jsonify(differences)

    except Exception as e:
        logger.error(f'Error comparing versions: {str(e)}')
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@bp.route('/versions', methods=['GET'])
def get_all_versions():
    """獲取所有版本列表"""
    try:
        # 獲取所有不重複的版本號，按版本號降序排序
        versions = db.session.query(MemberVersion.version)\
            .distinct()\
            .order_by(MemberVersion.version.desc())\
            .all()
        
        # 獲取每個版本的會員數量
        version_info = []
        for (version,) in versions:
            try:
                # 計算該版本的會員數量
                member_count = db.session.query(func.count(MemberVersion.id))\
                    .filter(MemberVersion.version == version)\
                    .scalar()
                
                # 獲取該版本的創建時間
                created_at = db.session.query(MemberVersion.created_at)\
                    .filter(MemberVersion.version == version)\
                    .order_by(MemberVersion.created_at.asc())\
                    .first()
                
                version_info.append({
                    'version': str(version),
                    'member_count': member_count,
                    'created_at': created_at[0].isoformat() if created_at else None
                })
            except Exception as e:
                logger.error(f'Error processing version {version}: {str(e)}')
                continue
        
        logger.info(f'Found {len(version_info)} versions')
        return jsonify(version_info)
        
    except Exception as e:
        logger.error(f'Error getting versions: {str(e)}')
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@bp.route('/compare-versions', methods=['POST'])
def compare_member_versions():
    """比較單一會員的兩個版本"""
    logger.info('Compare member versions request received')
    data = request.get_json()
    try:
        version1 = MemberVersion.query.filter_by(
            member_id=data['member_id'], 
            version=data['version1']
        ).first_or_404()
        version2 = MemberVersion.query.filter_by(
            member_id=data['member_id'], 
            version=data['version2']
        ).first_or_404()
        
        # Compare the two versions and return differences
        differences = {}
        fields = ['account', 'chinese_name', 'english_name', 'department_class',
                  'member_number', 'is_guest', 'is_admin', 'gender', 'handicap']
                  
        for field in fields:
            value1 = version1.data.get(field)
            value2 = version2.data.get(field)

            if value1 != value2:
                # 找到對應的會員姓名
                member_name = next(
                    (m.chinese_name for m in Member.query.filter_by(id=data['member_id']).all()),
                    None
                )

                differences[field] = {
                    'old': value1,
                    'new': value2
                }
        
        logger.info(f'Versions compared: {data["member_id"]}, version {version1.version} and version {version2.version}')
        return jsonify({
            'member_id': data['member_id'],
            'version1': version1.version,
            'version2': version2.version,
            'version1_date': version1.created_at,
            'version2_date': version2.created_at,
            'differences': differences
        })
    except Exception as e:
        logger.error(f'Error comparing versions: {str(e)}')
        logger.error(traceback.format_exc())
        return jsonify({'error': f'Error comparing versions: {str(e)}'}), 400

@bp.route('/versions/<version>', methods=['DELETE'])
def delete_version(version):
    """刪除指定版本的會員資料。只能刪除非最新版本。"""
    try:
        logger.info(f'收到刪除版本請求: {version}, 類型: {type(version)}')
        
        # 檢查是否為最新版本
        latest_version = db.session.query(MemberVersion.version)\
            .order_by(MemberVersion.version.desc())\
            .first()
            
        if not latest_version:
            logger.error('找不到任何版本')
            return jsonify({'error': '找不到任何版本'}), 404
            
        latest_version_str = str(latest_version[0])
        version_str = str(version)
        
        logger.info(f'最新版本: {latest_version_str}, 要刪除的版本: {version_str}')
        logger.info(f'版本比較結果: {version_str == latest_version_str}')
        
        # 檢查版本是否存在
        version_exists = db.session.query(MemberVersion)\
            .filter(MemberVersion.version == version_str)\
            .first()
            
        if not version_exists:
            logger.error(f'找不到要刪除的版本: {version_str}')
            return jsonify({'error': f'找不到版本 {version_str}'}), 404
            
        # 最新版本不能刪除，其他版本都可以刪除
        if str(latest_version[0]) == version_str:
            logger.error(f'嘗試刪除最新版本被拒絕: {version_str}')
            return jsonify({'error': '不能刪除最新版本'}), 400
        
        # 執行刪除
        logger.info(f'開始執行刪除操作: {version_str}')
        deleted_count = db.session.query(MemberVersion)\
            .filter(MemberVersion.version == version_str)\
            .delete()
            
        db.session.commit()
        logger.info(f'成功刪除版本 {version_str}，共刪除 {deleted_count} 筆記錄')
        return jsonify({'message': f'成功刪除版本 {version_str}'})
        
    except Exception as e:
        db.session.rollback()
        logger.error(f'刪除版本時發生異常: {str(e)}')
        logger.error(f'異常類型: {type(e)}')
        logger.error(f'完整異常追蹤: {traceback.format_exc()}')
        return jsonify({'error': str(e)}), 500

@bp.route('/clear', methods=['POST'])
def clear_members():
    try:
        # 先清除所有版本記錄
        MemberVersion.query.delete()
        # 再清除所有會員記錄
        Member.query.delete()
        db.session.commit()
        return jsonify({'message': '已清除所有會員資料和版本記錄'})
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error clearing members: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@bp.route('/count', methods=['GET'])
def get_member_count():
    """獲取最新版本的會員總數（不含來賓）"""
    try:
        # 獲取最新版本號
        latest_version_query = db.session.query(MemberVersion.version)\
            .order_by(MemberVersion.version.desc())\
            .first()
        
        if not latest_version_query:
            logger.warning("No versions found")
            return jsonify({'count': 0})
        
        latest_version = latest_version_query[0]    
        logger.info(f"Latest version found: {latest_version}")
        
        # 先獲取最新版本的所有會員資料
        members = db.session.query(Member)\
            .join(MemberVersion, Member.id == MemberVersion.member_id)\
            .filter(MemberVersion.version == latest_version)\
            .all()
        
        # 手動過濾非來賓會員
        non_guest_count = 0
        for member in members:
            # 獲取該會員在最新版本的資料
            version_data = db.session.query(MemberVersion)\
                .filter(
                    MemberVersion.member_id == member.id,
                    MemberVersion.version == latest_version
                ).first()
            
            if version_data:
                member_data = version_data.data
                is_guest = member_data.get('is_guest', False)
                if not is_guest:
                    non_guest_count += 1
                logger.debug(f"Member {member.id}: is_guest = {is_guest}")
            
        logger.info(f"Found {non_guest_count} non-guest members in version {latest_version}")
        return jsonify({'count': non_guest_count})
        
    except Exception as e:
        logger.error(f"獲取會員總數失敗: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': str(e)}), 500

@bp.route('/count', methods=['GET'])
def get_members_count():
    try:
        count = Member.query.count()
        return jsonify({'count': count})
    except Exception as e:
        current_app.logger.error(f"Error getting member count: {str(e)}")
        current_app.logger.error(traceback.format_exc())
        return jsonify({
            'error': 'Failed to get member count',
            'details': str(e)
        }), 500

@bp.route('/average-handicap', methods=['GET'])
def get_average_handicap():
    try:
        result = db.session.query(db.func.avg(Member.handicap)).filter(Member.handicap != None).scalar()
        average_handicap = float(result) if result is not None else 0
        return jsonify({'averageHandicap': round(average_handicap, 1)})
    except Exception as e:
        current_app.logger.error(f"Error calculating average handicap: {str(e)}")
        current_app.logger.error(traceback.format_exc())
        return jsonify({
            'error': 'Failed to calculate average handicap',
            'details': str(e)
        }), 500

@bp.route('/export', methods=['GET'])
def export_members():
    """匯出會員資料為 Excel 檔案"""
    try:
        # 從查詢參數中獲取版本號，如果沒有指定則使用最新版本
        version = request.args.get('version')
        
        # 獲取最新版本號
        latest_version = db.session.query(MemberVersion.version)\
            .order_by(MemberVersion.version.desc())\
            .first()
            
        if not version and latest_version:
            version = latest_version[0]
        elif not latest_version:
            logger.warning('No member versions found in database')
            return jsonify({'error': '沒有會員資料可供匯出'}), 400
            
        logger.info(f'Exporting members for version: {version}')
        
        # 使用指定版本號獲取會員資料，包括版本資料
        members_data = db.session.query(Member, MemberVersion)\
            .join(MemberVersion)\
            .filter(MemberVersion.version == version)\
            .order_by(Member.member_number)\
            .all()
            
        if not members_data:
            return jsonify({'error': f'版本 {version} 沒有會員資料'}), 404
            
        # 準備 Excel 數據
        data = []
        for member, version_data in members_data:
            # 從版本資料中獲取差點
            handicap = version_data.data.get('handicap') if version_data.data else member.handicap
            
            data.append({
                '會員編號': member.member_number,
                '帳號': member.account,
                '中文姓名': member.chinese_name,
                '英文姓名': member.english_name,
                '系級': member.department_class,
                '會員類型': '來賓' if member.is_guest else '會員',
                '管理權限': '是' if member.is_admin else '否',
                '差點': handicap
            })
            
        # 創建 DataFrame
        df = pd.DataFrame(data)
        
        # 創建一個 BytesIO 對象來保存 Excel 文件
        excel_file = io.BytesIO()
        
        # 將 DataFrame 寫入 Excel
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='會員資料')
            
        # 將指針移到文件開頭
        excel_file.seek(0)
        
        # 生成下載文件名
        current_time = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'members_v{version}_{current_time}.xlsx'
        
        # 返回 Excel 文件
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
            
    except Exception as e:
        logger.error(f'Error exporting members: {str(e)}')
        logger.error(traceback.format_exc())
        return jsonify({'error': '匯出會員資料失敗'}), 500

@bp.route('/template', methods=['GET'])
def download_template():
    """下載會員資料範本"""
    try:
        # 創建一個新的 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "會員資料"

        # 定義欄位順序（與上傳檔案格式一致）
        headers = [
            '會員類型',
            '帳號',
            '中文姓名',
            '英文姓名',
            '系級',
            '會員編號',
            '是否為管理員',
            '最新差點',
            '性別'
        ]

        # 設置欄位寬度
        for i, header in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = 15

        # 寫入標題
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # 設置樣式
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 為標題行添加邊框和填充
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.border = border
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # 使用 BytesIO 保存文件
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='member_template.xlsx'
        )

    except Exception as e:
        logger.error(f"生成範本時發生錯誤: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': '生成範本失敗'}), 500
